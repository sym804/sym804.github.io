"""Portfolio_QA_sym 의 issue_log.xlsx 신규 생성 + QA Strategy Playbook enhancement 항목 append.

글로벌 [최상위] 이슈 관리 원칙의 12 컬럼 양식을 따름.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import date
from pathlib import Path

OUT = Path(r'C:\Users\ymseo\Documents\Portfolio_QA_sym\issue_log.xlsx')

HEADERS = [
    '이슈번호', '발견일', '유형', '제목', '사전 조건',
    '재현 방법 / 개선 사항', '원인 및 수정 / 개선 사유', '영향 범위',
    '발생 버전', '수정 버전', '상태', '심각도',
]

ROW = [
    1,
    date(2026, 5, 2).isoformat(),
    'enhancement',
    'QA Engineering Strategy Playbook 추가',
    '',
    '풀스택 QA 전략 8섹션 멀티섹션 SPA 신규 자산. _qa-strategy/ (Vite + React + TS + Tailwind + sym-ui 컴포넌트 소스 복사), qa_strategy/ (정적 빌드 산출물). 3원칙 트라이앵글 (Risk-Based + Shift-Left + Automation ROI). 7 SVG 다이어그램. 4 케이스 섹션 Field Note 5 산출물 양방향 링크.',
    '외부 면접용 포트폴리오 + 사내 룰북 공통 자산 부재. sym-ui 디자인 시스템의 외부 프로젝트 사용 사례 첫 검증. typecheck / lint / build / WCAG 2.1 AA / visual regression baseline 7곳 모두 통과.',
    'Portfolio_QA_sym 폴더 (index.html 작품 그리드에 Project 06 카드 추가, _qa-strategy/ 소스, qa_strategy/ 빌드 산출물). 다른 산출물 페이지 영향 없음.',
    '',
    'Portfolio_QA_sym v0.1.0',
    '완료',
    'Major',
]

wb = Workbook()
ws = wb.active
ws.title = 'Issues'

# 헤더
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2A3E5C')
for col, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 컬럼 폭
widths = [10, 12, 14, 38, 16, 60, 60, 40, 18, 22, 10, 12]
for col, w in enumerate(widths, start=1):
    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

# 데이터 행
for col, v in enumerate(ROW, start=1):
    cell = ws.cell(row=2, column=col, value=v)
    cell.alignment = Alignment(vertical='top', wrap_text=True)

ws.row_dimensions[1].height = 26
ws.row_dimensions[2].height = 100

wb.save(OUT)
print(f'created: {OUT}')
