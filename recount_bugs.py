# -*- coding: utf-8 -*-
"""
'AI 로 제품을 만든다는 것'(ai_product.html) 집계 수치 재현/검증 스크립트.

문서 5절과 7절의 숫자를 stockradar 이슈로그 엑셀에서 규칙으로 산출한다.
숫자를 손으로 적지 않고 이 규칙으로 만들었음을 공개해, 값이 맞는지 검증할 수 있게 한다.

집계 규칙
- 버그 모집단 = 유형 칸이 명확히 'bug'/'버그' 인 행만 (초기 양식이 무너져 유형이
  소실된 행은 제외; 그 규모도 아래에서 함께 출력한다)
- 도메인 분류 = 제목/재현/원인/영향 텍스트를 대표 주제 하나로 '배타' 배정(한 버그=한 칸).
  우선순위 순서대로 첫 매칭 카테고리에 넣는다 -> 도메인 합 + 일반 합 = 전체 버그.

한계
- 배타 배정이라 순서에 의존한다. 순서와 무관하게 고정되는 값은 "도메인을 하나라도
  건드린 버그 수"(=도메인 우선 배정 시의 도메인 합)뿐이다. 개별 카테고리 분할은
  현재 정규식과 우선순위의 결과이지 '주 원인'의 분포가 아니다. 아래 견고성 블록에서
  이 민감도를 직접 출력한다.

실행: STOCKRADAR_DIR 환경변수로 stockradar 위치를 지정할 수 있다.
      기본값은 이 스크립트 기준 ../stockradar.
      python recount_bugs.py
"""
import openpyxl, sys, re, datetime, os
sys.stdout.reconfigure(encoding='utf-8')
from collections import Counter

HERE = os.path.dirname(os.path.abspath(__file__))
STOCKRADAR = os.environ.get('STOCKRADAR_DIR', os.path.join(HERE, '..', 'stockradar'))
KR = os.path.join(STOCKRADAR, 'web', 'issue_log.xlsx')
US = os.path.join(STOCKRADAR, 'web_us', 'issue_log.xlsx')


def load(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)
            if r[0] not in (None, '')]
    wb.close()
    return rows


kr, us = load(KR), load(US)
allrows = kr + us

# ── 전체 이슈 건수 ─────────────────────────────
print(f'[전체 이슈] KR={len(kr)}  US={len(us)}  합={len(allrows)}')

# ── 버그 모집단 (유형 명확) ────────────────────
BUG = {'bug', '버그'}
def is_bug(r):
    return str(r.get('유형', '')).strip().lower() in BUG
bugs = [r for r in allrows if is_bug(r)]
print(f'[버그 모집단] {len(bugs)} (KR {sum(is_bug(r) for r in kr)} / US {sum(is_bug(r) for r in us)})')


def text(r):
    return ' '.join(str(r.get(k, '') or '') for k in
                    ['제목', '재현 방법 / 개선 사항', '원인 및 수정 / 개선 사유', '영향 범위'])


# 배타 분류: 도메인 5 + 일반 5(기타 포함). 우선순위 = 리스트 순서.
CATS = [
    ('PIT/시점정합',  r'\bPIT\b|시점 ?정합|미래 ?참조|look.?ahead|룩어헤드|미확정 ?종가|당일.{0,6}(종가|매수|buy)|재무.{0,4}시점'),
    ('백테스트/퀀트', r'백테스트|back.?test|CAGR|샤프|리밸런싱|리밸런스|퀀트|팩터|시그널 ?점수|전략 ?점수|과대계상|수익률 ?계산'),
    ('ETF/배당',      r'ETF|배당|dividend|분배금|커버드콜|월배당|div_freq'),
    ('재무/공시',     r'\bPER\b|\bPBR\b|\bROE\b|\bEPS\b|\bBPS\b|재무|실적|공시|disclosure|시가총액|시총'),
    ('거래일/휴장',   r'거래일|휴장|휴일|영업일|장중|장마감|개장|폐장|상장폐지|거래정지|신고가|신저가|52주'),
    ('null/예외',     r'\bnull\b|None|예외|Exception|크래시|crash|500 ?에러|undefined|NaN|KeyError|traceback'),
    ('UI/레이아웃',   r'\bUI\b|레이아웃|정렬|색상|스타일|버튼|모바일|반응형|툴팁|렌더|표시 ?안|깨짐|차트'),
    ('성능',          r'성능|느림|지연|타임아웃|timeout|캐시|메모리|\bCPU\b|과부하|폴링'),
    ('API/계약',      r'\bAPI\b|엔드포인트|endpoint|응답 ?코드|스키마|파라미터|400|404|계약'),
]
DOMAIN = ['PIT/시점정합', '백테스트/퀀트', 'ETF/배당', '재무/공시', '거래일/휴장']

cnt = Counter()
for r in bugs:
    t = text(r)
    for name, pat in CATS:
        if re.search(pat, t, re.I):
            cnt[name] += 1
            break
    else:
        cnt['기타'] += 1

dsum = sum(cnt[c] for c in DOMAIN)
gsum = len(bugs) - dsum
print('[배타 분류]')
for c in DOMAIN:
    print(f'  도메인 {c}: {cnt[c]}')
for c in ['UI/레이아웃', '기타', 'null/예외', 'API/계약', '성능']:
    print(f'  일반  {c}: {cnt[c]}')
print(f'  => 도메인 합={dsum}  일반 합={gsum}  총={dsum+gsum}  도메인비율={dsum/len(bugs)*100:.1f}%')

# ── 7절: 세 칸(발견일/심각도/유형) 모두 해석되는 행 ──
VALID_TYPE = {'bug', '버그', 'enhancement', '개선', '기능', 'refactor', '문서', 'feature'}
VALID_SEV = {'block', 'critical', 'major', 'minor', 'trivial'}
def parse_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v
    s = str(v or '').strip()
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        try:
            return datetime.date(int(m[1]), int(m[2]), int(m[3]))
        except ValueError:
            return None
    return None
def interpretable(r):
    t = str(r.get('유형', '')).strip().lower() in VALID_TYPE
    s = str(r.get('심각도', '')).strip().lower() in VALID_SEV
    d = parse_date(r.get('발견일')) is not None
    return t and s and d
interp = [r for r in allrows if interpretable(r)]
print(f'[7절] 세 칸 해석 행: {len(interp)}/{len(allrows)} ({len(interp)/len(allrows)*100:.0f}%)')

def is_march(r):
    d = parse_date(r.get('발견일'))
    return d is not None and d.year == 2026 and d.month == 3
march_kr = [r for r in kr if str(r.get('발견일', '')).strip().startswith('2026-03')
            or is_march(r)]
march_interp = [r for r in march_kr if interpretable(r)]
print(f'[7절] 3월 KR 발견일표기 행 {len(march_kr)} 중 해석 {len(march_interp)}')

# ── 유형 분포: '365건 제외'의 근거 ─────────────────
ENH = {'enhancement', '개선', '기능', 'refactor', '문서', 'feature'}
def typ(r):
    t = str(r.get('유형', '')).strip().lower()
    return 'bug' if t in BUG else 'enh' if t in ENH else 'unknown'
tc = Counter(typ(r) for r in allrows)
print(f'[유형 분포] bug={tc["bug"]}  enh={tc["enh"]}  소실(unknown)={tc["unknown"]}'
      f'  <- 문서 "유형 소실 365건 제외" 근거')

# ── 견고성/민감도: 70%가 무엇에 의존하는가 ──
dom_pats = [pat for name, pat in CATS if name in DOMAIN]
gen_pats = [pat for name, pat in CATS if name not in DOMAIN]
def hit(r, pats):
    t = text(r)
    return any(re.search(p, t, re.I) for p in pats)
touch_dom = sum(1 for r in bugs if hit(r, dom_pats))   # 도메인 키워드 '접촉'
touch_gen = sum(1 for r in bugs if hit(r, gen_pats))
both = sum(1 for r in bugs if hit(r, dom_pats) and hit(r, gen_pats))
cnt_lo = Counter()
for r in bugs:
    t = text(r)
    for name, pat in [(n, p) for n, p in CATS if n not in DOMAIN] + \
                     [(n, p) for n, p in CATS if n in DOMAIN]:
        if re.search(pat, t, re.I):
            cnt_lo[name] += 1
            break
    else:
        cnt_lo['기타'] += 1
dom_lo = sum(cnt_lo[c] for c in DOMAIN)
print(f'[견고성] 도메인 접촉(순서 불변 상한)={touch_dom}  일반 접촉={touch_gen}  양쪽 동시={both}'
      f' ({both/len(bugs)*100:.0f}%)')
print(f'[견고성] 도메인 우선 배정={dsum}({dsum/len(bugs)*100:.1f}%)  '
      f'일반 우선 배정={dom_lo}({dom_lo/len(bugs)*100:.1f}%)')
print(f'  => "도메인을 하나라도 건드린 버그 {touch_dom}건"은 우선순위와 무관(=배타 도메인우선 {dsum}). '
      f'분할(도메인 {dsum} vs 일반 {len(bugs)-dsum})은 우선순위가 만든 것.')

# ── 제외한 365건이 비율에 준 방향 ──
excluded = [r for r in allrows if typ(r) == 'unknown']
ex_dom = sum(1 for r in excluded if hit(r, dom_pats))
mixed_ratio = (dsum + ex_dom) / (len(bugs) + len(excluded)) * 100
print(f'[제외 {len(excluded)} 방향] 제외분 도메인 접촉 {ex_dom}/{len(excluded)} ({ex_dom/len(excluded)*100:.0f}%). '
      f'전부 버그로 합치면 도메인비율 {dsum/len(bugs)*100:.1f}% -> {mixed_ratio:.1f}% 로 하락'
      f' (제외는 비율을 올리는 방향)')
